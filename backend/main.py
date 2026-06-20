import os
import re
import json
import base64
import tempfile
from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from typing import List, Optional
from pydantic import BaseModel
import pdfplumber
import anthropic
import openpyxl
from model_builder import build_consumer_retail_model

IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp", ".gif"}
IMAGE_MEDIA_TYPES = {
    ".png": "image/png",
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".webp": "image/webp",
    ".gif": "image/gif",
}

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

client = anthropic.Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY"))


# ── Keywords that identify the real consolidated financial statement pages ─────
FS_KEYWORDS = [
    "consolidated statements of income",
    "consolidated statements of operations",
    "consolidated balance sheet",
    "consolidated statements of cash flows",
    "consolidated statements of earnings",
    "statements of consolidated income",
]

# Keywords for pages to DEPRIORITISE (summaries, highlights, quarterly data)
NOISE_KEYWORDS = [
    "selected financial data",
    "quarterly financial data",
    "five-year",
    "five year",
    "highlights",
    "non-gaap",
]


def _page_content(page) -> str:
    """Extract tables + text from a single pdfplumber page."""
    parts = []
    tables = page.extract_tables()
    if tables:
        for t_idx, table in enumerate(tables):
            rows = [" | ".join(cell.strip().replace("\n", " ") if cell else "" for cell in row)
                    for row in table]
            parts.append(f"[Table {t_idx+1}]\n" + "\n".join(rows))
    text = page.extract_text()
    if text:
        parts.append(f"[Text]\n{text}")
    return "\n\n".join(parts)


def _score_page(content: str) -> int:
    """
    Score a page for relevance to consolidated financial statements.
    Higher = more relevant. Negative = noise to deprioritise.
    """
    lower = content.lower()
    score = 0
    for kw in FS_KEYWORDS:
        if kw in lower:
            score += 10
    for kw in NOISE_KEYWORDS:
        if kw in lower:
            score -= 5
    return score


def extract_content_from_pdfs(files: List[UploadFile]) -> str:
    all_sections = []
    for file in files:
        scored_pages = []
        with pdfplumber.open(file.file) as pdf:
            for page_num, page in enumerate(pdf.pages, 1):
                content = _page_content(page)
                if content.strip():
                    score = _score_page(content)
                    scored_pages.append((score, page_num, content))

        # Sort: financial statement pages first, noise pages last
        scored_pages.sort(key=lambda x: -x[0])

        file_sections = [f"=== FILE: {file.filename} ==="]
        for score, page_num, content in scored_pages:
            file_sections.append(f"-- Page {page_num} (relevance={score}) --\n{content}")

        all_sections.append("\n\n".join(file_sections))

    return "\n\n".join(all_sections)


def _is_image(filename: str) -> bool:
    return any(filename.lower().endswith(ext) for ext in IMAGE_EXTENSIONS)


def _media_type(filename: str) -> str:
    for ext, mt in IMAGE_MEDIA_TYPES.items():
        if filename.lower().endswith(ext):
            return mt
    return "image/png"


# ── Number parser (Python handles math, not Claude) ────────────────────────────
def parse_raw_number(raw: Optional[str], multiplier: float) -> Optional[float]:
    """Convert a raw string like '269,698' or '(4,580)' to a float in millions."""
    if raw is None or str(raw).strip() in ("", "null", "None", "—", "-"):
        return None
    s = str(raw).strip()
    is_negative = s.startswith("(") and s.endswith(")")
    s = re.sub(r"[^\d.]", "", s)
    if not s:
        return None
    value = float(s) * multiplier
    return -value if is_negative else value


UNIT_MULTIPLIERS = {
    "millions": 1.0,
    "thousands": 0.001,
    "billions": 1000.0,
}


def apply_multiplier(raw_dict: dict, multiplier: float) -> dict:
    """Recursively convert all raw string lists to float lists."""
    result = {}
    for k, v in raw_dict.items():
        if isinstance(v, list):
            result[k] = [parse_raw_number(item, multiplier) for item in v]
        elif isinstance(v, dict):
            result[k] = apply_multiplier(v, multiplier)
        else:
            result[k] = v
    return result


# ── Pass 1: Claude finds raw strings only ─────────────────────────────────────
def _extraction_prompt() -> str:
    return """You are extracting financial data from company filings. Your ONLY job is to find each line item and copy the exact text string as it appears — do not calculate, convert, round, or interpret anything.

RULES:
- Extract ONLY from the CONSOLIDATED STATEMENTS OF INCOME / OPERATIONS / EARNINGS — the primary audited financial statements. Ignore: selected financial data tables, quarterly summaries, 5-year summaries, highlights, non-GAAP tables, segment data, or any table that is not the main consolidated statement.
- Copy the exact characters from the filing including commas, dots, and parentheses for negatives e.g. (4,580)
- The filing shows numbers for multiple years in columns — return one value per year, oldest year first
- Use null if a line item is not found or not applicable
- Do NOT convert units — return numbers exactly as printed
- Do NOT compute derived values (do not calculate EBITDA, do not calculate FCF)
- "revenue" = net sales / merchandise sales ONLY, never includes membership fees
- "membership_fees" = membership fee revenue line item (separate from revenue)

First identify:
1. The CONSOLIDATED STATEMENTS OF INCOME (the main audited table, not summaries)
2. How many fiscal years are shown in that table (we want the 3 most recent)
3. The unit scale stated in that table's header (e.g. "in millions", "in thousands")

Return ONLY this JSON, no commentary:
{
  "company_name": "exact company name as stated",
  "ticker": "ticker symbol",
  "fiscal_year_end_month": "month name",
  "years": ["YYYY", "YYYY", "YYYY"],
  "unit_in_filing": "millions | thousands | billions",
  "income_statement": {
    "revenue": ["raw_string", "raw_string", "raw_string"],
    "membership_fees": ["raw_string_or_null", "raw_string_or_null", "raw_string_or_null"],
    "cogs": ["raw_string", "raw_string", "raw_string"],
    "gross_profit": ["raw_string_or_null", "raw_string_or_null", "raw_string_or_null"],
    "sga": ["raw_string", "raw_string", "raw_string"],
    "depreciation_amortization": ["raw_string_or_null", "raw_string_or_null", "raw_string_or_null"],
    "ebit": ["raw_string", "raw_string", "raw_string"],
    "interest_expense": ["raw_string_or_null", "raw_string_or_null", "raw_string_or_null"],
    "ebt": ["raw_string", "raw_string", "raw_string"],
    "tax": ["raw_string", "raw_string", "raw_string"],
    "net_income": ["raw_string", "raw_string", "raw_string"]
  },
  "balance_sheet": {
    "cash": ["raw_string", "raw_string", "raw_string"],
    "accounts_receivable": ["raw_string_or_null", "raw_string_or_null", "raw_string_or_null"],
    "inventory": ["raw_string", "raw_string", "raw_string"],
    "total_current_assets": ["raw_string", "raw_string", "raw_string"],
    "ppe_net": ["raw_string", "raw_string", "raw_string"],
    "total_assets": ["raw_string", "raw_string", "raw_string"],
    "accounts_payable": ["raw_string", "raw_string", "raw_string"],
    "total_current_liabilities": ["raw_string", "raw_string", "raw_string"],
    "total_debt": ["raw_string_or_null", "raw_string_or_null", "raw_string_or_null"],
    "total_equity": ["raw_string", "raw_string", "raw_string"]
  },
  "cash_flow": {
    "cfo": ["raw_string", "raw_string", "raw_string"],
    "capex": ["raw_string", "raw_string", "raw_string"],
    "dividends_paid": ["raw_string_or_null", "raw_string_or_null", "raw_string_or_null"]
  },
  "operating_metrics": {
    "store_count": ["raw_string_or_null", "raw_string_or_null", "raw_string_or_null"],
    "comp_sales_growth": ["raw_string_or_null", "raw_string_or_null", "raw_string_or_null"],
    "membership_count": ["raw_string_or_null", "raw_string_or_null", "raw_string_or_null"],
    "membership_renewal_rate": ["raw_string_or_null", "raw_string_or_null", "raw_string_or_null"]
  },
  "shares_outstanding": ["raw_string_or_null", "raw_string_or_null", "raw_string_or_null"],
  "key_value_drivers": [
    {"driver": "name", "description": "why it matters for valuation", "raw_current_value": "raw_string", "unit": "unit string"},
    {"driver": "name", "description": "why it matters for valuation", "raw_current_value": "raw_string", "unit": "unit string"},
    {"driver": "name", "description": "why it matters for valuation", "raw_current_value": "raw_string", "unit": "unit string"}
  ]
}"""


def extract_raw_strings_from_text(raw_content: str) -> dict:
    """PDF path: send extracted text to Claude."""
    prompt = f"{_extraction_prompt()}\n\nFiling content:\n{raw_content[:100000]}"
    message = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=4096,
        temperature=0,
        messages=[{"role": "user", "content": prompt}],
    )
    text = message.content[0].text
    return json.loads(text[text.find("{"):text.rfind("}") + 1])


def extract_raw_strings_from_images(image_files: List[UploadFile]) -> dict:
    """Image path: send images directly to Claude vision — most accurate method."""
    content = []
    for f in image_files:
        raw_bytes = f.file.read()
        b64 = base64.standard_b64encode(raw_bytes).decode("utf-8")
        content.append({
            "type": "image",
            "source": {"type": "base64", "media_type": _media_type(f.filename), "data": b64},
        })
    content.append({"type": "text", "text": _extraction_prompt()})

    message = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=4096,
        temperature=0,
        messages=[{"role": "user", "content": content}],
    )
    text = message.content[0].text
    return json.loads(text[text.find("{"):text.rfind("}") + 1])


# ── Pass 2: Python converts strings → floats ───────────────────────────────────
def build_financials(raw: dict) -> dict:
    unit = raw.get("unit_in_filing", "millions").lower()
    multiplier = UNIT_MULTIPLIERS.get(unit, 1.0)

    is_raw = raw["income_statement"]
    bs_raw = raw["balance_sheet"]
    cf_raw = raw["cash_flow"]
    op_raw = raw["operating_metrics"]

    def nums(section, key):
        return [parse_raw_number(v, multiplier) for v in section.get(key, [None, None, None])]

    def pct(section, key):
        # Percentages: strip % sign, divide by 100
        vals = []
        for v in section.get(key, [None, None, None]):
            if v is None or str(v).strip() in ("", "null", "None"):
                vals.append(None)
                continue
            s = str(v).replace("%", "").strip()
            n = parse_raw_number(s, 1.0)
            vals.append(n)  # store as plain number, e.g. 5.1 for 5.1%
        return vals

    revenue = nums(is_raw, "revenue")
    cogs    = nums(is_raw, "cogs")
    gross_p = nums(is_raw, "gross_profit")
    sga     = nums(is_raw, "sga")
    da      = nums(is_raw, "depreciation_amortization")
    ebit    = nums(is_raw, "ebit")

    # Compute EBITDA in Python: EBIT + D&A (never ask Claude to compute this)
    ebitda = []
    for i in range(3):
        e = ebit[i]
        d = da[i]
        ebitda.append((e + d) if (e is not None and d is not None) else None)

    cfo    = nums(cf_raw, "cfo")
    capex  = nums(cf_raw, "capex")
    # FCF in Python: CFO + Capex (capex is negative)
    fcf = []
    for i in range(3):
        c = cfo[i]
        x = capex[i]
        fcf.append((c + x) if (c is not None and x is not None) else None)

    # Process key value drivers
    kvd_raw = raw.get("key_value_drivers", [])
    kvd = []
    for d in kvd_raw[:3]:
        raw_val = d.get("raw_current_value")
        parsed  = parse_raw_number(str(raw_val).replace("%", ""), 1.0) if raw_val else None
        kvd.append({
            "driver": d.get("driver", ""),
            "description": d.get("description", ""),
            "current_value": parsed,
            "unit": d.get("unit", ""),
        })

    return {
        "company_name": raw.get("company_name", ""),
        "ticker": raw.get("ticker", ""),
        "currency": "USD",
        "unit_in_filing": unit,
        "fiscal_year_end": raw.get("fiscal_year_end_month", ""),
        "years": raw.get("years", []),
        "income_statement": {
            "revenue":                   revenue,
            "membership_fees":           nums(is_raw, "membership_fees"),
            "cogs":                      cogs,
            "gross_profit":              gross_p,
            "sga":                       sga,
            "ebitda":                    ebitda,
            "depreciation_amortization": da,
            "ebit":                      ebit,
            "interest_expense":          nums(is_raw, "interest_expense"),
            "ebt":                       nums(is_raw, "ebt"),
            "tax":                       nums(is_raw, "tax"),
            "net_income":                nums(is_raw, "net_income"),
        },
        "balance_sheet": {
            "cash":                      nums(bs_raw, "cash"),
            "accounts_receivable":       nums(bs_raw, "accounts_receivable"),
            "inventory":                 nums(bs_raw, "inventory"),
            "total_current_assets":      nums(bs_raw, "total_current_assets"),
            "ppe_net":                   nums(bs_raw, "ppe_net"),
            "total_assets":              nums(bs_raw, "total_assets"),
            "accounts_payable":          nums(bs_raw, "accounts_payable"),
            "total_current_liabilities": nums(bs_raw, "total_current_liabilities"),
            "total_debt":                nums(bs_raw, "total_debt"),
            "total_equity":              nums(bs_raw, "total_equity"),
        },
        "cash_flow": {
            "cfo":            cfo,
            "capex":          capex,
            "fcf":            fcf,
            "dividends_paid": nums(cf_raw, "dividends_paid"),
        },
        "operating_metrics": {
            "store_count":             nums(op_raw, "store_count"),
            "comp_sales_growth":       pct(op_raw, "comp_sales_growth"),
            "membership_count":        nums(op_raw, "membership_count"),
            "membership_renewal_rate": pct(op_raw, "membership_renewal_rate"),
        },
        "shares_outstanding": nums(raw, "shares_outstanding"),
        "key_value_drivers":  kvd,
        "_raw_strings":       raw,
        "_computed":          [],   # filled by fill_gaps()
    }


# ── Pass 3: auto-derive + suggest ambiguous items ─────────────────────────────
def _get(lst, i):
    try:
        return lst[i]
    except (IndexError, TypeError):
        return None

def _fill(lst, i, value, label, computed: list):
    if _get(lst, i) is None and value is not None:
        lst[i] = round(value, 1)
        computed.append(f"{label} [year {i+1}]")

def _compute_values(fin: dict, terms: list) -> list:
    """
    Compute per-year values from a list of (section, field, sign[, optional]) terms.
    optional=True means treat None as 0 rather than blocking the whole formula.
    Returns list of floats or Nones.
    """
    n = len(fin.get("years", []))
    results = []
    for i in range(n):
        total = 0.0
        ok = True
        for term in terms:
            section, field, sign = term[0], term[1], term[2]
            optional = term[3] if len(term) > 3 else False
            v = _get(fin[section].get(field, []), i)
            if v is None:
                if optional:
                    v = 0.0   # missing optional term → treat as zero
                else:
                    ok = False
                    break
            total += v * sign
        results.append(round(total, 1) if ok else None)
    return results


# Rules for items that need analyst approval when missing.
# Each rule: field, section, label, list of formula suggestions.
# Each suggestion: display string, description, terms list [(section, field, sign)]
PENDING_RULES = [
    {
        "field": "gross_profit",
        "section": "income_statement",
        "label": "Gross Profit",
        "suggestions": [
            {
                "id": "gp_total_rev",
                "formula": "Total Revenue − Merchandise Costs",
                "description": "Includes membership fees — matches Costco's P&L structure where total revenue is the top line",
                "terms": [
                    ("income_statement", "revenue", 1),
                    ("income_statement", "membership_fees", 1, True),  # optional: 0 if missing
                    ("income_statement", "cogs", -1),
                ],
            },
            {
                "id": "gp_net_sales",
                "formula": "Net Sales − Merchandise Costs",
                "description": "Excludes membership fees — treats membership as other income",
                "terms": [
                    ("income_statement", "revenue", 1),
                    ("income_statement", "cogs", -1),
                ],
            },
        ],
    },
    {
        "field": "ebit",
        "section": "income_statement",
        "label": "EBIT / Operating Income",
        "suggestions": [
            {
                "id": "ebit_from_gp",
                "formula": "Gross Profit − SG&A",
                "description": "Standard derivation once gross profit is known",
                "terms": [
                    ("income_statement", "gross_profit", 1),
                    ("income_statement", "sga", -1),
                ],
            },
            {
                "id": "ebit_from_rev",
                "formula": "Net Sales − COGS − SG&A",
                "description": "Direct from revenue without going through gross profit",
                "terms": [
                    ("income_statement", "revenue", 1),
                    ("income_statement", "cogs", -1),
                    ("income_statement", "sga", -1),
                ],
            },
        ],
    },
    {
        "field": "ebt",
        "section": "income_statement",
        "label": "Earnings Before Tax",
        "suggestions": [
            {
                "id": "ebt_ebit_int",
                "formula": "EBIT + Interest Expense",
                "description": "Interest expense is stored as negative so this nets it out",
                "terms": [
                    ("income_statement", "ebit", 1),
                    ("income_statement", "interest_expense", 1),
                ],
            },
        ],
    },
    {
        "field": "net_income",
        "section": "income_statement",
        "label": "Net Income",
        "suggestions": [
            {
                "id": "ni_ebt_tax",
                "formula": "EBT − Tax",
                "description": "Standard bottom-line derivation",
                "terms": [
                    ("income_statement", "ebt", 1),
                    ("income_statement", "tax", -1),
                ],
            },
        ],
    },
    {
        "field": "fcf",
        "section": "cash_flow",
        "label": "Free Cash Flow",
        "suggestions": [
            {
                "id": "fcf_cfo_capex",
                "formula": "Cash from Operations + Capex",
                "description": "Capex is stored as negative so this subtracts it",
                "terms": [
                    ("cash_flow", "cfo", 1),
                    ("cash_flow", "capex", 1),
                ],
            },
        ],
    },
]


def auto_derive(fin: dict):
    """Auto-apply unambiguous identities that don't need analyst approval."""
    is_  = fin["income_statement"]
    bs   = fin["balance_sheet"]
    cf   = fin["cash_flow"]
    comp = fin.setdefault("_computed", [])
    n    = len(fin.get("years", []))

    for i in range(n):
        ebit   = _get(is_["ebit"], i)
        da     = _get(is_["depreciation_amortization"], i)
        ebitda = _get(is_["ebitda"], i)
        cfo    = _get(cf["cfo"], i)
        capex  = _get(cf["capex"], i)
        fcf    = _get(cf["fcf"], i)
        cash   = _get(bs["cash"], i)
        ar     = _get(bs["accounts_receivable"], i)
        inv    = _get(bs["inventory"], i)
        tca    = _get(bs["total_current_assets"], i)
        cogs   = _get(is_["cogs"], i)
        rev    = _get(is_["revenue"], i)
        gp     = _get(is_["gross_profit"], i)

        # EBITDA = EBIT + D&A  (unambiguous)
        if ebitda is None and ebit is not None and da is not None:
            _fill(is_["ebitda"], i, ebit + da, "EBITDA = EBIT + D&A", comp)

        # COGS reverse
        if cogs is None and rev is not None and gp is not None:
            _fill(is_["cogs"], i, rev - gp, "COGS = Revenue − Gross Profit", comp)

        # Total Current Assets
        if tca is None and cash is not None and inv is not None:
            _fill(bs["total_current_assets"], i,
                  (cash or 0) + (ar or 0) + inv,
                  "Total Current Assets ≈ Cash + AR + Inventory", comp)


def build_pending_derivations(fin: dict) -> list:
    """
    For each rule in PENDING_RULES, check if the field is null.
    If so, compute preview values for each suggestion and return them.
    Only suggestions where ALL component fields are available are included.
    """
    pending = []
    for rule in PENDING_RULES:
        section = rule["section"]
        field   = rule["field"]
        vals    = fin[section].get(field, [])
        n       = len(fin.get("years", []))

        if all(_get(vals, i) is not None for i in range(n)):
            continue   # already populated — skip

        valid_suggestions = []
        for sug in rule["suggestions"]:
            preview = _compute_values(fin, sug["terms"])
            if any(v is not None for v in preview):
                valid_suggestions.append({**sug, "preview_values": preview})

        if valid_suggestions:
            pending.append({
                "field":       field,
                "section":     section,
                "label":       rule["label"],
                "suggestions": valid_suggestions,
            })

    return pending


# ── Endpoints ──────────────────────────────────────────────────────────────────
@app.post("/api/extract-preview")
async def extract_preview(files: List[UploadFile] = File(...)):
    if not files:
        raise HTTPException(status_code=400, detail="No files uploaded")
    try:
        image_files = [f for f in files if _is_image(f.filename or "")]
        pdf_files   = [f for f in files if not _is_image(f.filename or "")]

        if image_files and not pdf_files:
            raw_strings = extract_raw_strings_from_images(image_files)
        elif pdf_files:
            raw_content = extract_content_from_pdfs(pdf_files)
            raw_strings = extract_raw_strings_from_text(raw_content)
        else:
            raise HTTPException(status_code=400, detail="No valid files. Upload PDFs or images (PNG/JPG).")

        financials = build_financials(raw_strings)
        auto_derive(financials)
        pending = build_pending_derivations(financials)
        return {"financials": financials, "pending_derivations": pending}
    except json.JSONDecodeError:
        raise HTTPException(status_code=500, detail="Failed to parse AI response. Try uploading cleaner files.")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


class BuildModelRequest(BaseModel):
    financials: dict

@app.post("/api/build-model")
async def build_model(req: BuildModelRequest):
    try:
        output_path = tempfile.mktemp(suffix=".xlsx")
        build_consumer_retail_model(req.financials, output_path)
        company = req.financials.get("company_name", "Company").replace(" ", "_")
        return FileResponse(
            output_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=f"{company}_Model.xlsx"
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── Excel chat ─────────────────────────────────────────────────────────────────
def parse_excel_to_context(file) -> dict:
    wb = openpyxl.load_workbook(file, data_only=True)
    sheets = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            if any(c is not None for c in row):
                rows.append([str(c) if c is not None else "" for c in row])
        sheets[name] = rows
    return sheets


def build_excel_context_string(sheets: dict) -> str:
    parts = []
    for name, rows in sheets.items():
        parts.append(f"=== SHEET: {name} ===")
        for row in rows:
            parts.append("\t".join(row))
    return "\n".join(parts)


class ChatMessage(BaseModel):
    role: str
    content: str

class ChatRequest(BaseModel):
    excel_context: str
    messages: List[ChatMessage]


@app.post("/api/parse-excel")
async def parse_excel(file: UploadFile = File(...)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only Excel files are supported")
    try:
        sheets   = parse_excel_to_context(file.file)
        context  = build_excel_context_string(sheets)
        return {
            "context":     context[:80000],
            "sheet_names": list(sheets.keys()),
            "row_counts":  {s: len(r) for s, r in sheets.items()},
            "filename":    file.filename,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/chat")
async def chat(req: ChatRequest):
    system = f"""You are an expert equity analyst assistant. The user has uploaded a financial model in Excel.
Answer questions accurately and insightfully — interpreting numbers, explaining assumptions, flagging risks.
Ground every answer in the actual model data. If something isn't in the model, say so.
Be concise but complete. Use numbers when relevant.

Excel model content:
{req.excel_context}"""

    response = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=1024,
        temperature=0,
        system=system,
        messages=[{"role": m.role, "content": m.content} for m in req.messages],
    )
    return {"reply": response.content[0].text}


@app.get("/health")
def health():
    return {"status": "ok"}
