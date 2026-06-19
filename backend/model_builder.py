from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
import openpyxl.styles.numbers as num_formats


def _r(num, den, default):
    """Safe ratio: returns default if either operand is None/zero."""
    if num is None or den is None or den == 0:
        return default
    return num / den

def _v(val, default=0):
    """Return val if not None, else default."""
    return val if val is not None else default


# ── Colour palette ────────────────────────────────────────────────────────────
DARK_BLUE  = "1F3864"
MID_BLUE   = "2E75B6"
LIGHT_BLUE = "D6E4F0"
GREY_BG    = "F2F2F2"
GREEN_BG   = "E2EFDA"
ORANGE_BG  = "FCE4D6"
YELLOW_BG  = "FFF2CC"
WHITE      = "FFFFFF"

# ── Number formats ────────────────────────────────────────────────────────────
FMT_COMMA0 = '#,##0'
FMT_COMMA1 = '#,##0.0'
FMT_COMMA2 = '#,##0.00'
FMT_PCT1   = '0.0%'
FMT_PCT2   = '0.00%'
FMT_MULT   = '0.0x'
FMT_DEC2   = '0.00'


# ── Helpers ───────────────────────────────────────────────────────────────────
def _font(bold=False, color=WHITE, size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def _bottom_border():
    s = Side(style="thin")
    return Border(bottom=s)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _pct(value):
    """Return None-safe percentage (already as decimal for Excel)."""
    if value is None:
        return None
    return value / 100.0

def _safe(lst, idx, default=None):
    try:
        v = lst[idx]
        return v if v is not None else default
    except (IndexError, TypeError):
        return default

def _col(c):
    return get_column_letter(c)


def _header_row(ws, row, label, col_start, col_end, bg=DARK_BLUE):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start, value=label)
    cell.font = _font(bold=True, size=11)
    cell.fill = _fill(bg)
    cell.alignment = _align("left")


def _section_header(ws, row, label, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = _font(bold=True, color=WHITE, size=10)
    cell.fill = _fill(MID_BLUE)
    cell.alignment = _align("left")


def _label(ws, row, col, text, indent=0, bold=False, italic=False, color="000000"):
    cell = ws.cell(row=row, column=col, value="  " * indent + text)
    cell.font = Font(bold=bold, italic=italic, color=color, size=10)
    cell.alignment = _align("left")


def _value(ws, row, col, val, fmt=FMT_COMMA1, bold=False):
    cell = ws.cell(row=row, column=col, value=val)
    cell.number_format = fmt
    cell.font = Font(bold=bold, size=10)
    cell.alignment = _align("right")
    return cell


def _formula(ws, row, col, formula, fmt=FMT_COMMA1, bold=False, fill_hex=None):
    cell = ws.cell(row=row, column=col, value=formula)
    cell.number_format = fmt
    cell.font = Font(bold=bold, size=10)
    cell.alignment = _align("right")
    if fill_hex:
        cell.fill = _fill(fill_hex)
    return cell


def _assumption(ws, row, col, val, fmt=FMT_PCT1):
    """Blue input cell for assumptions."""
    cell = ws.cell(row=row, column=col, value=val)
    cell.number_format = fmt
    cell.font = Font(color="1F497D", bold=True, size=10)
    cell.fill = _fill(LIGHT_BLUE)
    cell.alignment = _align("right")
    return cell


def _set_col_widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ═════════════════════════════════════════════════════════════════════════════
# SHEET 1 – ASSUMPTIONS
# ═════════════════════════════════════════════════════════════════════════════
def _build_assumptions(wb, fin):
    ws = wb.create_sheet("Assumptions")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {"A": 38, "B": 16, "C": 16, "D": 16, "E": 16, "F": 16})

    years = fin["years"]
    proj_years = [str(int(years[-1]) + i) for i in range(1, 6)]
    all_years = years + proj_years
    H = len(all_years)

    # Title
    ws.row_dimensions[1].height = 28
    _header_row(ws, 1, "ASSUMPTIONS & DRIVERS", 1, H + 1)

    # Year headers
    ws.row_dimensions[2].height = 18
    ws.cell(row=2, column=1, value="").font = _font(bold=True, color="000000")
    for i, y in enumerate(all_years):
        c = ws.cell(row=2, column=i + 2, value=y)
        c.font = Font(bold=True, size=10)
        c.alignment = _align("center")
        c.fill = _fill(GREY_BG if i < len(years) else LIGHT_BLUE)

    def _hist_label(col):
        return col <= len(years) + 1  # +1 for label column

    row = 3
    # ── Revenue ─────────────────────────────────────────────────────────────
    _section_header(ws, row, "REVENUE ASSUMPTIONS", H + 1); row += 1

    rev = fin["income_statement"]["revenue"]
    comp = fin["operating_metrics"].get("comp_sales_growth", [None]*3)
    store = fin["operating_metrics"].get("store_count", [None]*3)

    # Historical comp sales growth
    _label(ws, row, 1, "Comp Sales Growth")
    for i, v in enumerate(comp):
        _value(ws, row, i + 2, _pct(v), FMT_PCT1)
    # Projected comp growth (declining slightly)
    base = (_pct(comp[-1]) if comp[-1] else 0.05)
    proj_comps = [max(base - 0.005 * j, 0.03) for j in range(5)]
    for j, v in enumerate(proj_comps):
        _assumption(ws, row, len(years) + 2 + j, v, FMT_PCT1)
    ws.cell(row=row, column=1).font = Font(size=10)
    # Store the assumption row for reference
    comp_row = row; row += 1

    _label(ws, row, 1, "New Store Openings")
    for i, v in enumerate(store):
        _value(ws, row, i + 2, v, FMT_COMMA0)
    for j in range(5):
        _assumption(ws, row, len(years) + 2 + j, 25, FMT_COMMA0)
    store_opens_row = row; row += 1

    _label(ws, row, 1, "Revenue Growth Rate (YoY)")
    for i in range(len(years)):
        if i == 0:
            _value(ws, row, i + 2, None, FMT_PCT1)
        else:
            r_col = _col(i + 2)
            r_prev = _col(i + 1)
            _formula(ws, row, i + 2, f"=Income_Statement!{r_col}3/Income_Statement!{r_prev}3-1", FMT_PCT1)
    for j in range(5):
        c = len(years) + 2 + j
        _formula(ws, row, c, f"=Assumptions!{_col(c)}5*0.6+Assumptions!{_col(c-1 if c > len(years)+2 else c)}5*0.4", FMT_PCT1, fill_hex=GREEN_BG)
    rev_growth_row = row; row += 1

    row += 1  # spacer
    # ── Margins ──────────────────────────────────────────────────────────────
    _section_header(ws, row, "MARGIN ASSUMPTIONS", H + 1); row += 1

    gm = fin["income_statement"]
    _label(ws, row, 1, "Gross Margin")
    for i in range(len(years)):
        r_col = _col(i + 2)
        _formula(ws, row, i + 2, f"=Income_Statement!{r_col}5/Income_Statement!{r_col}3", FMT_PCT1)
    base_gm = _r(gm["gross_profit"][-1], gm["revenue"][-1], 0.13)
    for j in range(5):
        _assumption(ws, row, len(years) + 2 + j, base_gm + 0.001 * j, FMT_PCT1)
    gm_row = row; row += 1

    _label(ws, row, 1, "SG&A as % of Revenue")
    for i in range(len(years)):
        r_col = _col(i + 2)
        _formula(ws, row, i + 2, f"=Income_Statement!{r_col}7/Income_Statement!{r_col}3", FMT_PCT1)
    base_sga = _r(gm["sga"][-1], gm["revenue"][-1], 0.10)
    for j in range(5):
        _assumption(ws, row, len(years) + 2 + j, base_sga - 0.001 * j, FMT_PCT1)
    sga_row = row; row += 1

    _label(ws, row, 1, "EBITDA Margin")
    for i in range(len(years)):
        r_col = _col(i + 2)
        _formula(ws, row, i + 2, f"=Income_Statement!{r_col}9/Income_Statement!{r_col}3", FMT_PCT1)
    for j in range(5):
        c = len(years) + 2 + j
        gm_ref = f"Assumptions!{_col(c)}{gm_row}"
        sga_ref = f"Assumptions!{_col(c)}{sga_row}"
        _formula(ws, row, c, f"={gm_ref}-{sga_ref}", FMT_PCT1, fill_hex=GREEN_BG)
    ebitda_margin_row = row; row += 1

    row += 1
    # ── Membership ───────────────────────────────────────────────────────────
    mem_fees = fin["income_statement"].get("membership_fees", [None, None, None])
    if any(v for v in mem_fees if v):
        _section_header(ws, row, "MEMBERSHIP ASSUMPTIONS", H + 1); row += 1

        mem_count = fin["operating_metrics"].get("membership_count", [None]*3)
        mem_renew  = fin["operating_metrics"].get("membership_renewal_rate", [None]*3)

        _label(ws, row, 1, "Membership Count (millions)")
        for i, v in enumerate(mem_count):
            _value(ws, row, i + 2, v, FMT_COMMA1)
        base_mem = mem_count[-1] or 70
        for j in range(5):
            _assumption(ws, row, len(years) + 2 + j, base_mem * (1.06 ** (j + 1)), FMT_COMMA1)
        mem_count_row = row; row += 1

        _label(ws, row, 1, "Renewal Rate")
        for i, v in enumerate(mem_renew):
            _value(ws, row, i + 2, _pct(v), FMT_PCT1)
        base_renew = _pct(mem_renew[-1]) if mem_renew[-1] else 0.925
        for j in range(5):
            _assumption(ws, row, len(years) + 2 + j, min(base_renew + 0.002 * j, 0.95), FMT_PCT1)
        mem_renew_row = row; row += 1

        _label(ws, row, 1, "Avg Membership Fee (USD)")
        avg_fee = _r(mem_fees[-1], mem_count[-1], 65)
        for i, v in enumerate(mem_fees):
            cnt = mem_count[i] or 1
            _value(ws, row, i + 2, (v / cnt) if v else None, FMT_COMMA2)
        for j in range(5):
            _assumption(ws, row, len(years) + 2 + j, avg_fee * (1.05 ** (j + 1)), FMT_COMMA2)
        mem_fee_row = row; row += 1

    row += 1
    # ── Capex & D&A ──────────────────────────────────────────────────────────
    _section_header(ws, row, "CAPEX & DEPRECIATION ASSUMPTIONS", H + 1); row += 1

    capex = fin["cash_flow"]["capex"]
    da    = fin["income_statement"]["depreciation_amortization"]
    rev_vals = fin["income_statement"]["revenue"]

    _label(ws, row, 1, "Capex as % of Revenue")
    for i in range(len(years)):
        r_col = _col(i + 2)
        _formula(ws, row, i + 2, f"=ABS(Cash_Flow!{r_col}4)/Income_Statement!{r_col}3", FMT_PCT1)
    base_capex = _r(abs(capex[-1]) if capex[-1] is not None else None, rev_vals[-1], 0.025)
    for j in range(5):
        _assumption(ws, row, len(years) + 2 + j, base_capex, FMT_PCT1)
    capex_pct_row = row; row += 1

    _label(ws, row, 1, "D&A as % of Revenue")
    for i in range(len(years)):
        r_col = _col(i + 2)
        _formula(ws, row, i + 2, f"=Income_Statement!{r_col}11/Income_Statement!{r_col}3", FMT_PCT1)
    base_da = _r(da[-1], rev_vals[-1], 0.01)
    for j in range(5):
        _assumption(ws, row, len(years) + 2 + j, base_da, FMT_PCT1)
    da_pct_row = row; row += 1

    row += 1
    # ── Working Capital ──────────────────────────────────────────────────────
    _section_header(ws, row, "WORKING CAPITAL ASSUMPTIONS", H + 1); row += 1

    inv   = fin["balance_sheet"]["inventory"]
    ar    = fin["balance_sheet"]["accounts_receivable"]
    ap    = fin["balance_sheet"]["accounts_payable"]

    _label(ws, row, 1, "Days Inventory Outstanding")
    for i in range(len(years)):
        inv_v = inv[i] or 0; rev_v = rev_vals[i] or 1
        _value(ws, row, i + 2, inv_v / rev_v * 365, FMT_DEC2)
    base_dio = _r(inv[-1], rev_vals[-1], 30 / 365) * 365
    for j in range(5):
        _assumption(ws, row, len(years) + 2 + j, base_dio, FMT_DEC2)
    dio_row = row; row += 1

    _label(ws, row, 1, "Days Payable Outstanding")
    for i in range(len(years)):
        ap_v = ap[i] or 0; rev_v = rev_vals[i] or 1
        _value(ws, row, i + 2, ap_v / rev_v * 365, FMT_DEC2)
    base_dpo = _r(ap[-1], rev_vals[-1], 30 / 365) * 365
    for j in range(5):
        _assumption(ws, row, len(years) + 2 + j, base_dpo, FMT_DEC2)
    dpo_row = row; row += 1

    row += 1
    # ── Valuation ────────────────────────────────────────────────────────────
    _section_header(ws, row, "VALUATION ASSUMPTIONS", H + 1); row += 1

    _label(ws, row, 1, "WACC")
    _assumption(ws, row, 2, 0.085, FMT_PCT1)
    wacc_cell = f"Assumptions!B{row}"; wacc_row = row; row += 1

    _label(ws, row, 1, "Terminal Growth Rate")
    _assumption(ws, row, 2, 0.03, FMT_PCT1)
    tgr_cell = f"Assumptions!B{row}"; tgr_row = row; row += 1

    _label(ws, row, 1, "EV/EBITDA Exit Multiple")
    _assumption(ws, row, 2, 20.0, FMT_COMMA1)
    exit_mult_row = row; row += 1

    _label(ws, row, 1, "Tax Rate")
    tax_vals = fin["income_statement"]["tax"]
    ebt_vals  = fin["income_statement"]["ebt"]
    base_tax = _r(abs(tax_vals[-1]) if tax_vals[-1] is not None else None, ebt_vals[-1], 0.25)
    _assumption(ws, row, 2, base_tax, FMT_PCT1)
    tax_row = row; row += 1

    # Store references for other sheets
    ws._assumption_rows = {
        "comp_row": comp_row,
        "rev_growth_row": rev_growth_row,
        "gm_row": gm_row,
        "sga_row": sga_row,
        "ebitda_margin_row": ebitda_margin_row,
        "capex_pct_row": capex_pct_row,
        "da_pct_row": da_pct_row,
        "dio_row": dio_row,
        "dpo_row": dpo_row,
        "wacc_row": wacc_row,
        "tgr_row": tgr_row,
        "exit_mult_row": exit_mult_row,
        "tax_row": tax_row,
        "proj_years": proj_years,
        "years": years,
    }
    return ws


# ═════════════════════════════════════════════════════════════════════════════
# SHEET 2 – INCOME STATEMENT
# ═════════════════════════════════════════════════════════════════════════════
def _build_income_statement(wb, fin, assumptions_ws):
    ws = wb.create_sheet("Income_Statement")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {"A": 36, "B": 16, "C": 16, "D": 16, "E": 16,
                         "F": 16, "G": 16, "H": 16, "I": 16})

    ar  = assumptions_ws._assumption_rows
    years = ar["years"]
    proj  = ar["proj_years"]
    all_y = years + proj
    H = len(all_y)

    inc = fin["income_statement"]

    ws.row_dimensions[1].height = 28
    _header_row(ws, 1, "INCOME STATEMENT  (USD millions)", 1, H + 1)

    ws.row_dimensions[2].height = 18
    for i, y in enumerate(all_y):
        c = ws.cell(row=2, column=i + 2, value=y)
        c.font = Font(bold=True, size=10)
        c.alignment = _align("center")
        c.fill = _fill(GREY_BG if i < len(years) else LIGHT_BLUE)

    def hval(row_num, lst, idx):
        _value(ws, row_num, idx + 2, _safe(lst, idx), FMT_COMMA1)

    def proj_col(j):
        return len(years) + 2 + j

    row = 3

    # Revenue
    _label(ws, row, 1, "Net Revenue", bold=True)
    for i in range(len(years)):
        hval(row, inc["revenue"], i)
    for j in range(5):
        c = proj_col(j)
        prev_rev = f"{_col(c-1)}{row}"
        rg_ref   = f"Assumptions!{_col(c)}{ar['rev_growth_row']}"
        _formula(ws, row, c, f"={prev_rev}*(1+{rg_ref})", FMT_COMMA1, bold=True, fill_hex=GREEN_BG)
    rev_row = row; row += 1

    # Membership fees
    mem_fees = inc.get("membership_fees", [None]*3)
    if any(v for v in mem_fees if v):
        _label(ws, row, 1, "  Membership Fee Revenue", indent=1)
        for i in range(len(years)):
            hval(row, mem_fees, i)
        for j in range(5):
            c = proj_col(j)
            # mem count × renewal rate × avg fee (from assumptions)
            _formula(ws, row, c, f"=Assumptions!{_col(c)}{ar.get('mem_fee_row',row)}*1", FMT_COMMA1, fill_hex=GREEN_BG)
        mem_row = row; row += 1

    # COGS
    _label(ws, row, 1, "Cost of Goods Sold")
    for i in range(len(years)):
        hval(row, inc["cogs"], i)
    for j in range(5):
        c = proj_col(j)
        gm_ref  = f"Assumptions!{_col(c)}{ar['gm_row']}"
        rev_ref = f"{_col(c)}{rev_row}"
        _formula(ws, row, c, f"={rev_ref}*(1-{gm_ref})", FMT_COMMA1)
    cogs_row = row; row += 1

    # Gross Profit
    _label(ws, row, 1, "Gross Profit", bold=True)
    for i in range(len(years)):
        hval(row, inc["gross_profit"], i)
    for j in range(5):
        c = proj_col(j)
        _formula(ws, row, c, f"={_col(c)}{rev_row}-{_col(c)}{cogs_row}", FMT_COMMA1, bold=True, fill_hex=GREEN_BG)
    gp_row = row

    # Gross Margin %
    row += 1
    _label(ws, row, 1, "  Gross Margin %", indent=1, italic=True)
    for j, col_idx in enumerate(range(2, H + 2)):
        _formula(ws, row, col_idx, f"={_col(col_idx)}{gp_row}/{_col(col_idx)}{rev_row}", FMT_PCT1)
    row += 1

    # SG&A
    _label(ws, row, 1, "SG&A Expenses")
    for i in range(len(years)):
        hval(row, inc["sga"], i)
    for j in range(5):
        c = proj_col(j)
        sga_ref = f"Assumptions!{_col(c)}{ar['sga_row']}"
        rev_ref = f"{_col(c)}{rev_row}"
        _formula(ws, row, c, f"={rev_ref}*{sga_ref}", FMT_COMMA1)
    sga_row = row; row += 1

    # EBITDA
    _label(ws, row, 1, "EBITDA", bold=True)
    for i in range(len(years)):
        hval(row, inc["ebitda"], i)
    for j in range(5):
        c = proj_col(j)
        _formula(ws, row, c, f"={_col(c)}{gp_row}-{_col(c)}{sga_row}", FMT_COMMA1, bold=True, fill_hex=GREEN_BG)
    ebitda_row = row

    row += 1
    _label(ws, row, 1, "  EBITDA Margin %", indent=1, italic=True)
    for col_idx in range(2, H + 2):
        _formula(ws, row, col_idx, f"={_col(col_idx)}{ebitda_row}/{_col(col_idx)}{rev_row}", FMT_PCT1)
    row += 1

    # D&A
    _label(ws, row, 1, "Depreciation & Amortization")
    for i in range(len(years)):
        hval(row, inc["depreciation_amortization"], i)
    for j in range(5):
        c = proj_col(j)
        da_ref  = f"Assumptions!{_col(c)}{ar['da_pct_row']}"
        rev_ref = f"{_col(c)}{rev_row}"
        _formula(ws, row, c, f"={rev_ref}*{da_ref}", FMT_COMMA1)
    da_row = row; row += 1

    # EBIT
    _label(ws, row, 1, "EBIT (Operating Income)", bold=True)
    for i in range(len(years)):
        hval(row, inc["ebit"], i)
    for j in range(5):
        c = proj_col(j)
        _formula(ws, row, c, f"={_col(c)}{ebitda_row}-{_col(c)}{da_row}", FMT_COMMA1, bold=True, fill_hex=GREEN_BG)
    ebit_row = row; row += 1

    # Interest
    _label(ws, row, 1, "Interest Expense")
    for i in range(len(years)):
        hval(row, inc["interest_expense"], i)
    base_int = inc["interest_expense"][-1] or 0
    for j in range(5):
        _value(ws, row, proj_col(j), base_int, FMT_COMMA1)
    int_row = row; row += 1

    # EBT
    _label(ws, row, 1, "Earnings Before Tax", bold=True)
    for i in range(len(years)):
        hval(row, inc["ebt"], i)
    for j in range(5):
        c = proj_col(j)
        _formula(ws, row, c, f"={_col(c)}{ebit_row}-{_col(c)}{int_row}", FMT_COMMA1, bold=True)
    ebt_row = row; row += 1

    # Tax
    _label(ws, row, 1, "Income Tax")
    for i in range(len(years)):
        hval(row, inc["tax"], i)
    for j in range(5):
        c = proj_col(j)
        _formula(ws, row, c, f"={_col(c)}{ebt_row}*Assumptions!B{ar['tax_row']}", FMT_COMMA1)
    tax_row = row; row += 1

    # Net Income
    _label(ws, row, 1, "Net Income", bold=True)
    for i in range(len(years)):
        hval(row, inc["net_income"], i)
    for j in range(5):
        c = proj_col(j)
        _formula(ws, row, c, f"={_col(c)}{ebt_row}-{_col(c)}{tax_row}", FMT_COMMA1, bold=True, fill_hex=GREEN_BG)
    ni_row = row; row += 1

    # Net Margin %
    _label(ws, row, 1, "  Net Margin %", indent=1, italic=True)
    for col_idx in range(2, H + 2):
        _formula(ws, row, col_idx, f"={_col(col_idx)}{ni_row}/{_col(col_idx)}{rev_row}", FMT_PCT1)
    row += 1

    # EPS
    shares = fin.get("shares_outstanding", [None]*3)
    _label(ws, row, 1, "Shares Outstanding (millions)")
    for i, v in enumerate(shares):
        _value(ws, row, i + 2, v, FMT_COMMA1)
    base_shr = shares[-1] or 440
    for j in range(5):
        _value(ws, row, proj_col(j), base_shr * 0.99, FMT_COMMA1)
    shr_row = row; row += 1

    _label(ws, row, 1, "EPS (Diluted)", bold=True)
    for i in range(len(years)):
        shr = _safe(shares, i) or 1
        ni  = _safe(inc["net_income"], i) or 0
        _value(ws, row, i + 2, ni / shr, FMT_COMMA2, bold=True)
    for j in range(5):
        c = proj_col(j)
        _formula(ws, row, c, f"={_col(c)}{ni_row}/{_col(c)}{shr_row}", FMT_COMMA2, bold=True, fill_hex=GREEN_BG)

    # Store row references for DCF sheet
    ws._rows = {
        "rev_row": rev_row,
        "ebitda_row": ebitda_row,
        "ebit_row": ebit_row,
        "da_row": da_row,
        "ni_row": ni_row,
        "shr_row": shr_row,
    }
    return ws


# ═════════════════════════════════════════════════════════════════════════════
# SHEET 3 – BALANCE SHEET
# ═════════════════════════════════════════════════════════════════════════════
def _build_balance_sheet(wb, fin, assumptions_ws, is_ws):
    ws = wb.create_sheet("Balance_Sheet")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {"A": 36, "B": 16, "C": 16, "D": 16, "E": 16,
                         "F": 16, "G": 16, "H": 16, "I": 16})

    ar   = assumptions_ws._assumption_rows
    ir   = is_ws._rows
    years = ar["years"]
    proj  = ar["proj_years"]
    all_y = years + proj
    H = len(all_y)
    bs = fin["balance_sheet"]

    _header_row(ws, 1, "BALANCE SHEET  (USD millions)", 1, H + 1)
    ws.row_dimensions[1].height = 28
    for i, y in enumerate(all_y):
        c = ws.cell(row=2, column=i + 2, value=y)
        c.font = Font(bold=True, size=10)
        c.alignment = _align("center")
        c.fill = _fill(GREY_BG if i < len(years) else LIGHT_BLUE)

    def hval(r, lst, idx):
        _value(ws, r, idx + 2, _safe(lst, idx), FMT_COMMA1)

    def pj(j):
        return len(years) + 2 + j

    row = 3
    _section_header(ws, row, "ASSETS", H + 1); row += 1

    _label(ws, row, 1, "Cash & Equivalents")
    for i in range(len(years)):
        hval(row, bs["cash"], i)
    base_cash = bs["cash"][-1] or 10000
    for j in range(5):
        _formula(ws, row, pj(j), f"=Cash_Flow!{_col(pj(j))}6+Balance_Sheet!{_col(pj(j)-1)}{row}", FMT_COMMA1)
    cash_row = row; row += 1

    _label(ws, row, 1, "Accounts Receivable")
    for i in range(len(years)):
        hval(row, bs["accounts_receivable"], i)
    for j in range(5):
        c = pj(j)
        _formula(ws, row, c, f"=Income_Statement!{_col(c)}{ir['rev_row']}/365*30", FMT_COMMA1)
    ar_row = row; row += 1

    _label(ws, row, 1, "Inventory")
    for i in range(len(years)):
        hval(row, bs["inventory"], i)
    for j in range(5):
        c = pj(j)
        dio_ref = f"Assumptions!{_col(c)}{ar['dio_row']}"
        rev_ref = f"Income_Statement!{_col(c)}{ir['rev_row']}"
        _formula(ws, row, c, f"={rev_ref}/365*{dio_ref}", FMT_COMMA1)
    inv_row = row; row += 1

    _label(ws, row, 1, "Total Current Assets", bold=True)
    for i in range(len(years)):
        hval(row, bs["total_current_assets"], i)
    for j in range(5):
        c = pj(j)
        _formula(ws, row, c, f"={_col(c)}{cash_row}+{_col(c)}{ar_row}+{_col(c)}{inv_row}", FMT_COMMA1, bold=True, fill_hex=GREEN_BG)
    tca_row = row; row += 1

    _label(ws, row, 1, "PP&E, Net")
    for i in range(len(years)):
        hval(row, bs["ppe_net"], i)
    for j in range(5):
        c = pj(j)
        capex_ref = f"ABS(Cash_Flow!{_col(c)}4)"
        da_ref    = f"Income_Statement!{_col(c)}{ir['da_row']}"
        _formula(ws, row, c, f"={_col(c-1)}{row}+{capex_ref}-{da_ref}", FMT_COMMA1)
    ppe_row = row; row += 1

    _label(ws, row, 1, "Total Assets", bold=True)
    for i in range(len(years)):
        hval(row, bs["total_assets"], i)
    for j in range(5):
        c = pj(j)
        _formula(ws, row, c, f"={_col(c)}{tca_row}+{_col(c)}{ppe_row}", FMT_COMMA1, bold=True, fill_hex=GREEN_BG)
    ta_row = row; row += 1

    row += 1
    _section_header(ws, row, "LIABILITIES & EQUITY", H + 1); row += 1

    _label(ws, row, 1, "Accounts Payable")
    for i in range(len(years)):
        hval(row, bs["accounts_payable"], i)
    for j in range(5):
        c = pj(j)
        dpo_ref = f"Assumptions!{_col(c)}{ar['dpo_row']}"
        rev_ref = f"Income_Statement!{_col(c)}{ir['rev_row']}"
        _formula(ws, row, c, f"={rev_ref}/365*{dpo_ref}", FMT_COMMA1)
    ap_row = row; row += 1

    _label(ws, row, 1, "Total Current Liabilities", bold=True)
    for i in range(len(years)):
        hval(row, bs["total_current_liabilities"], i)
    for j in range(5):
        c = pj(j)
        _formula(ws, row, c, f"={_col(c)}{ap_row}*1.4", FMT_COMMA1, bold=True)
    tcl_row = row; row += 1

    _label(ws, row, 1, "Total Debt")
    for i in range(len(years)):
        hval(row, bs["total_debt"], i)
    base_debt = bs["total_debt"][-1] or 0
    for j in range(5):
        _value(ws, row, pj(j), base_debt, FMT_COMMA1)
    debt_row = row; row += 1

    _label(ws, row, 1, "Total Equity", bold=True)
    for i in range(len(years)):
        hval(row, bs["total_equity"], i)
    for j in range(5):
        c = pj(j)
        ni_ref  = f"Income_Statement!{_col(c)}{ir['ni_row']}"
        div_ref = f"ABS(Cash_Flow!{_col(c)}5)"
        _formula(ws, row, c, f"={_col(c-1)}{row}+{ni_ref}-{div_ref}", FMT_COMMA1, bold=True, fill_hex=GREEN_BG)
    eq_row = row; row += 1

    ws._rows = {"debt_row": debt_row, "eq_row": eq_row, "cash_row": cash_row}
    return ws


# ═════════════════════════════════════════════════════════════════════════════
# SHEET 4 – CASH FLOW
# ═════════════════════════════════════════════════════════════════════════════
def _build_cash_flow(wb, fin, assumptions_ws, is_ws):
    ws = wb.create_sheet("Cash_Flow")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {"A": 36, "B": 16, "C": 16, "D": 16, "E": 16,
                         "F": 16, "G": 16, "H": 16, "I": 16})

    ar   = assumptions_ws._assumption_rows
    ir   = is_ws._rows
    years = ar["years"]
    proj  = ar["proj_years"]
    all_y = years + proj
    H = len(all_y)
    cf = fin["cash_flow"]

    _header_row(ws, 1, "CASH FLOW STATEMENT  (USD millions)", 1, H + 1)
    ws.row_dimensions[1].height = 28
    for i, y in enumerate(all_y):
        c = ws.cell(row=2, column=i + 2, value=y)
        c.font = Font(bold=True, size=10)
        c.alignment = _align("center")
        c.fill = _fill(GREY_BG if i < len(years) else LIGHT_BLUE)

    def hval(r, lst, idx):
        _value(ws, r, idx + 2, _safe(lst, idx), FMT_COMMA1)

    def pj(j):
        return len(years) + 2 + j

    row = 3
    # CFO
    _label(ws, row, 1, "Cash from Operations", bold=True)
    for i in range(len(years)):
        hval(row, cf["cfo"], i)
    for j in range(5):
        c = pj(j)
        ebitda_ref = f"Income_Statement!{_col(c)}{ir['ebitda_row']}"
        _formula(ws, row, c, f"={ebitda_ref}*0.90", FMT_COMMA1, bold=True, fill_hex=GREEN_BG)
    cfo_row = row; row += 1

    # Capex
    _label(ws, row, 1, "Capital Expenditures")
    for i in range(len(years)):
        hval(row, cf["capex"], i)
    for j in range(5):
        c = pj(j)
        capex_ref = f"Assumptions!{_col(c)}{ar['capex_pct_row']}"
        rev_ref   = f"Income_Statement!{_col(c)}{ir['rev_row']}"
        _formula(ws, row, c, f"=-{rev_ref}*{capex_ref}", FMT_COMMA1)
    capex_row = row; row += 1

    # FCF
    _label(ws, row, 1, "Free Cash Flow", bold=True)
    for i in range(len(years)):
        hval(row, cf["fcf"], i)
    for j in range(5):
        c = pj(j)
        _formula(ws, row, c, f"={_col(c)}{cfo_row}+{_col(c)}{capex_row}", FMT_COMMA1, bold=True, fill_hex=GREEN_BG)
    fcf_row = row; row += 1

    # Dividends
    _label(ws, row, 1, "Dividends Paid")
    for i in range(len(years)):
        hval(row, cf["dividends_paid"], i)
    base_div = cf["dividends_paid"][-1] or 0
    for j in range(5):
        _value(ws, row, pj(j), -(abs(base_div) * (1.05 ** (j + 1))), FMT_COMMA1)
    div_row = row; row += 1

    # Net change in cash
    _label(ws, row, 1, "Net Change in Cash", bold=True)
    for j in range(5):
        c = pj(j)
        _formula(ws, row, c, f"={_col(c)}{fcf_row}+{_col(c)}{div_row}", FMT_COMMA1, bold=True, fill_hex=GREEN_BG)
    net_cash_row = row; row += 1

    ws._rows = {"fcf_row": fcf_row, "capex_row": capex_row, "div_row": div_row, "net_cash_row": net_cash_row}
    return ws


# ═════════════════════════════════════════════════════════════════════════════
# SHEET 5 – DCF VALUATION
# ═════════════════════════════════════════════════════════════════════════════
def _build_dcf(wb, fin, assumptions_ws, is_ws, bs_ws, cf_ws):
    ws = wb.create_sheet("DCF_Valuation")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {"A": 36, "B": 16, "C": 16, "D": 16, "E": 16,
                         "F": 16, "G": 16, "H": 16})

    ar   = assumptions_ws._assumption_rows
    ir   = is_ws._rows
    bsr  = bs_ws._rows
    cfr  = cf_ws._rows
    proj  = ar["proj_years"]
    years = ar["years"]
    H5   = len(proj)

    _header_row(ws, 1, "DCF VALUATION  (USD millions)", 1, H5 + 2)
    ws.row_dimensions[1].height = 28

    for j, y in enumerate(proj):
        c = ws.cell(row=2, column=j + 2, value=y)
        c.font = Font(bold=True, size=10)
        c.alignment = _align("center")
        c.fill = _fill(LIGHT_BLUE)

    base_col = len(years) + 2  # first projection column in other sheets

    def is_col(j):  # column in Income_Statement for proj year j
        return _col(base_col + j)

    def cf_col(j):
        return _col(base_col + j)

    row = 3
    _section_header(ws, row, "FREE CASH FLOW TO FIRM (FCFF)", H5 + 1); row += 1

    # EBIT
    _label(ws, row, 1, "EBIT")
    for j in range(H5):
        _formula(ws, row, j + 2, f"=Income_Statement!{is_col(j)}{ir['ebit_row']}", FMT_COMMA1)
    ebit_row = row; row += 1

    # NOPAT
    _label(ws, row, 1, "NOPAT  (EBIT × (1-tax))")
    for j in range(H5):
        _formula(ws, row, j + 2,
                 f"={_col(j+2)}{ebit_row}*(1-Assumptions!B{ar['tax_row']})", FMT_COMMA1)
    nopat_row = row; row += 1

    # D&A add-back
    _label(ws, row, 1, "+ Depreciation & Amortization")
    for j in range(H5):
        _formula(ws, row, j + 2, f"=Income_Statement!{is_col(j)}{ir['da_row']}", FMT_COMMA1)
    da_row = row; row += 1

    # Capex
    _label(ws, row, 1, "- Capital Expenditures")
    for j in range(H5):
        _formula(ws, row, j + 2, f"=ABS(Cash_Flow!{cf_col(j)}{cfr['capex_row']})", FMT_COMMA1)
    capex_row = row; row += 1

    # ΔWC
    _label(ws, row, 1, "- Change in Working Capital")
    for j in range(H5):
        _formula(ws, row, j + 2, f"=Income_Statement!{is_col(j)}{ir['rev_row']}*0.01", FMT_COMMA1)
    wc_row = row; row += 1

    # FCFF
    _label(ws, row, 1, "Free Cash Flow to Firm (FCFF)", bold=True)
    for j in range(H5):
        _formula(ws, row, j + 2,
                 f"={_col(j+2)}{nopat_row}+{_col(j+2)}{da_row}-{_col(j+2)}{capex_row}-{_col(j+2)}{wc_row}",
                 FMT_COMMA1, bold=True, fill_hex=GREEN_BG)
    fcff_row = row; row += 1

    # Discount factors
    _label(ws, row, 1, "Discount Factor  (WACC)")
    for j in range(H5):
        _formula(ws, row, j + 2,
                 f"=1/(1+Assumptions!B{ar['wacc_row']})^{j+1}", FMT_COMMA2)
    disc_row = row; row += 1

    # PV of FCFF
    _label(ws, row, 1, "PV of FCFF", bold=True)
    for j in range(H5):
        _formula(ws, row, j + 2,
                 f"={_col(j+2)}{fcff_row}*{_col(j+2)}{disc_row}",
                 FMT_COMMA1, bold=True, fill_hex=GREEN_BG)
    pv_fcff_row = row; row += 1

    row += 1
    _section_header(ws, row, "TERMINAL VALUE", H5 + 1); row += 1

    # TV (Gordon Growth)
    _label(ws, row, 1, "Terminal Value (Gordon Growth)")
    _formula(ws, row, 2,
             f"={_col(H5+1)}{fcff_row}*(1+Assumptions!B{ar['tgr_row']})/(Assumptions!B{ar['wacc_row']}-Assumptions!B{ar['tgr_row']})",
             FMT_COMMA1, bold=True)
    tv_gg_row = row; row += 1

    # TV (Exit Multiple)
    _label(ws, row, 1, "Terminal Value (EV/EBITDA Exit Multiple)")
    _formula(ws, row, 2,
             f"=Income_Statement!{is_col(H5-1)}{ir['ebitda_row']}*Assumptions!B{ar['exit_mult_row']}",
             FMT_COMMA1, bold=True)
    tv_em_row = row; row += 1

    # PV of TV
    _label(ws, row, 1, "PV of Terminal Value (Gordon Growth)")
    _formula(ws, row, 2,
             f"=B{tv_gg_row}/(1+Assumptions!B{ar['wacc_row']})^{H5}",
             FMT_COMMA1, bold=True, fill_hex=GREEN_BG)
    pv_tv_gg_row = row; row += 1

    _label(ws, row, 1, "PV of Terminal Value (Exit Multiple)")
    _formula(ws, row, 2,
             f"=B{tv_em_row}/(1+Assumptions!B{ar['wacc_row']})^{H5}",
             FMT_COMMA1, bold=True, fill_hex=GREEN_BG)
    pv_tv_em_row = row; row += 1

    row += 1
    _section_header(ws, row, "BRIDGE TO EQUITY VALUE", H5 + 1); row += 1

    # Sum PV of FCFFs
    _label(ws, row, 1, "Sum of PV(FCFF)")
    _formula(ws, row, 2, f"=SUM(B{pv_fcff_row}:{_col(H5+1)}{pv_fcff_row})", FMT_COMMA1)
    sum_pv_row = row; row += 1

    # Enterprise Values
    _label(ws, row, 1, "Enterprise Value — Gordon Growth", bold=True)
    _formula(ws, row, 2, f"=B{sum_pv_row}+B{pv_tv_gg_row}", FMT_COMMA1, bold=True, fill_hex=GREEN_BG)
    ev_gg_row = row; row += 1

    _label(ws, row, 1, "Enterprise Value — Exit Multiple", bold=True)
    _formula(ws, row, 2, f"=B{sum_pv_row}+B{pv_tv_em_row}", FMT_COMMA1, bold=True, fill_hex=GREEN_BG)
    ev_em_row = row; row += 1

    last_bs_col = _col(len(years) + 1)

    _label(ws, row, 1, "(-) Net Debt")
    _formula(ws, row, 2,
             f"=Balance_Sheet!{last_bs_col}{bsr['debt_row']}-Balance_Sheet!{last_bs_col}{bsr['cash_row']}",
             FMT_COMMA1)
    net_debt_row = row; row += 1

    last_is_col = _col(len(years) + 1)

    _label(ws, row, 1, "Shares Outstanding (millions)")
    _formula(ws, row, 2,
             f"=Income_Statement!{last_is_col}{ir['shr_row']}", FMT_COMMA1)
    shr_dcf_row = row; row += 1

    row += 1
    _section_header(ws, row, "TARGET PRICE", H5 + 1); row += 1

    _label(ws, row, 1, "Implied Share Price — Gordon Growth", bold=True)
    _formula(ws, row, 2,
             f"=(B{ev_gg_row}-B{net_debt_row})/B{shr_dcf_row}",
             FMT_COMMA2, bold=True, fill_hex=YELLOW_BG)
    tp_gg_row = row; row += 1

    _label(ws, row, 1, "Implied Share Price — Exit Multiple", bold=True)
    _formula(ws, row, 2,
             f"=(B{ev_em_row}-B{net_debt_row})/B{shr_dcf_row}",
             FMT_COMMA2, bold=True, fill_hex=YELLOW_BG)
    tp_em_row = row; row += 1

    ws._rows = {"tp_gg_row": tp_gg_row, "tp_em_row": tp_em_row, "ev_gg_row": ev_gg_row}
    return ws


# ═════════════════════════════════════════════════════════════════════════════
# SHEET 6 – SENSITIVITY ANALYSIS
# ═════════════════════════════════════════════════════════════════════════════
def _build_sensitivity(wb, fin, assumptions_ws, dcf_ws):
    ws = wb.create_sheet("Sensitivity_Analysis")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {"A": 28, "B": 14, "C": 14, "D": 14, "E": 14, "F": 14,
                         "G": 14, "H": 14, "I": 14, "J": 14, "K": 14})

    ar  = assumptions_ws._assumption_rows
    dr  = dcf_ws._rows

    drivers = fin.get("key_value_drivers", [])
    while len(drivers) < 3:
        drivers.append({"driver": f"Driver {len(drivers)+1}", "description": "", "current_value": 0, "unit": "%"})

    _header_row(ws, 1, "SENSITIVITY & SCENARIO ANALYSIS", 1, 11)
    ws.row_dimensions[1].height = 28

    row = 3
    ws.cell(row=row, column=1, value="This sheet shows how the DCF-implied share price changes when the 3 key value drivers vary.")
    ws.cell(row=row, column=1).font = Font(italic=True, size=9, color="555555")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
    row += 2

    # ── Build one sensitivity table per driver pair ───────────────────────
    def _sensitivity_table(start_row, title, row_label, row_ref_cell,
                           row_vals, col_label, col_ref_cell, col_vals, output_formula):
        r = start_row
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(col_vals)+2)
        t = ws.cell(row=r, column=1, value=title)
        t.font = Font(bold=True, size=11, color=WHITE)
        t.fill = _fill(DARK_BLUE)
        t.alignment = _align("center")
        r += 1

        # Axis labels
        ws.cell(row=r, column=1, value=f"↓ {row_label}  /  {col_label} →")
        ws.cell(row=r, column=1).font = Font(bold=True, size=9)
        for k, cv in enumerate(col_vals):
            c = ws.cell(row=r, column=k + 3, value=cv)
            c.font = Font(bold=True, size=10)
            c.fill = _fill(MID_BLUE)
            c.font = Font(bold=True, color=WHITE, size=10)
            c.alignment = _align("center")
        r += 1

        for rv in row_vals:
            rcel = ws.cell(row=r, column=2, value=rv)
            rcel.font = Font(bold=True, color=WHITE, size=10)
            rcel.fill = _fill(MID_BLUE)
            rcel.alignment = _align("center")
            for k, cv in enumerate(col_vals):
                formula = output_formula(rv, cv, row_ref_cell, col_ref_cell)
                cell = ws.cell(row=r, column=k + 3, value=formula)
                cell.number_format = FMT_COMMA2
                cell.alignment = _align("center")
                cell.fill = _fill(GREEN_BG)
                cell.font = Font(size=10)
            r += 1

        return r + 1

    tp_ref  = f"DCF_Valuation!B{dr['tp_gg_row']}"
    wacc_ref = f"Assumptions!B{ar['wacc_row']}"
    tgr_ref  = f"Assumptions!B{ar['tgr_row']}"
    gm_ref   = f"Assumptions!{_col(len(ar['years'])+2)}{ar['gm_row']}"
    comp_ref = f"Assumptions!{_col(len(ar['years'])+2)}{ar['comp_row']}"
    em_ref   = f"Assumptions!B{ar['exit_mult_row']}"

    def wacc_tgr_formula(rv, cv, rref, cref):
        return (f"=({tp_ref}*({wacc_ref}-{rv})/({wacc_ref}-{tgr_ref}))"
                f"*(1+({cv}-{tgr_ref}))/1")

    # Table 1: WACC vs Terminal Growth Rate
    wacc_range = [0.065, 0.075, 0.085, 0.095, 0.105]
    tgr_range  = [0.015, 0.020, 0.025, 0.030, 0.035]

    def sensitivity_formula_wacc_tgr(rv, cv, rref, cref):
        # Approximate: Price scales inversely with WACC-g spread
        return (f"={tp_ref}*({wacc_ref}-{tgr_ref})/({rv}-{cv})")

    row = _sensitivity_table(
        row,
        f"TABLE 1 — Target Price vs WACC & Terminal Growth Rate  |  Base: ${tp_ref}",
        "WACC", wacc_ref, wacc_range,
        "Terminal Growth Rate", tgr_ref, tgr_range,
        sensitivity_formula_wacc_tgr
    )

    # Table 2: Revenue Growth vs EBITDA Margin
    rev_growth_range = [0.04, 0.06, 0.08, 0.10, 0.12]
    ebitda_margin_range = [0.04, 0.05, 0.06, 0.07, 0.08]

    ebitda_m_ref = f"Assumptions!{_col(len(ar['years'])+2)}{ar['ebitda_margin_row']}"
    rev_g_ref    = f"Assumptions!{_col(len(ar['years'])+2)}{ar['rev_growth_row']}"

    def sensitivity_formula_rev_ebitda(rv, cv, rref, cref):
        return (f"={tp_ref}*({rev_g_ref}/{rv})*({cv}/{ebitda_m_ref})")

    row = _sensitivity_table(
        row,
        "TABLE 2 — Target Price vs Revenue Growth & EBITDA Margin",
        "Revenue Growth (Y1)", rev_g_ref, rev_growth_range,
        "EBITDA Margin", ebitda_m_ref, ebitda_margin_range,
        sensitivity_formula_rev_ebitda
    )

    # Table 3: EV/EBITDA Exit Multiple vs Revenue Growth
    exit_mult_range = [14, 16, 18, 20, 22, 24]
    comp_growth_range = [0.04, 0.05, 0.06, 0.07, 0.08]

    tp_em_ref = f"DCF_Valuation!B{dr['tp_em_row']}"

    def sensitivity_formula_mult_comp(rv, cv, rref, cref):
        return (f"={tp_em_ref}*({rv}/{em_ref})*({cv}/{comp_ref})")

    row = _sensitivity_table(
        row,
        "TABLE 3 — Target Price vs EV/EBITDA Exit Multiple & Comp Sales Growth",
        "EV/EBITDA Multiple", em_ref, exit_mult_range,
        "Comp Sales Growth", comp_ref, comp_growth_range,
        sensitivity_formula_mult_comp
    )

    # ── Scenario Analysis ────────────────────────────────────────────────────
    row += 1
    _header_row(ws, row, "SCENARIO ANALYSIS", 1, 6)
    ws.row_dimensions[row].height = 22
    row += 1

    scenarios = {
        "Bull Case": {"rev_growth": 0.12, "ebitda_margin": 0.07, "wacc": 0.075, "tgr": 0.035, "mult": 24},
        "Base Case": {"rev_growth": 0.08, "ebitda_margin": 0.055, "wacc": 0.085, "tgr": 0.030, "mult": 20},
        "Bear Case": {"rev_growth": 0.04, "ebitda_margin": 0.04, "wacc": 0.095, "tgr": 0.020, "mult": 16},
    }

    headers = ["Scenario", "Rev Growth", "EBITDA Margin", "WACC", "Term. Growth", "Exit Multiple", "Approx. Target Price"]
    for k, h in enumerate(headers):
        c = ws.cell(row=row, column=k + 1, value=h)
        c.font = Font(bold=True, color=WHITE, size=10)
        c.fill = _fill(MID_BLUE)
        c.alignment = _align("center")
    row += 1

    fill_map = {"Bull Case": GREEN_BG, "Base Case": YELLOW_BG, "Bear Case": ORANGE_BG}
    for name, vals in scenarios.items():
        ws.cell(row=row, column=1, value=name).font = Font(bold=True, size=10)
        ws.cell(row=row, column=1).fill = _fill(fill_map[name])

        ws.cell(row=row, column=2, value=vals["rev_growth"]).number_format = FMT_PCT1
        ws.cell(row=row, column=3, value=vals["ebitda_margin"]).number_format = FMT_PCT1
        ws.cell(row=row, column=4, value=vals["wacc"]).number_format = FMT_PCT1
        ws.cell(row=row, column=5, value=vals["tgr"]).number_format = FMT_PCT1
        ws.cell(row=row, column=6, value=vals["mult"]).number_format = FMT_COMMA1

        # Approximate price: DCF × adjustments
        approx = (f"={tp_ref}"
                  f"*({vals['rev_growth']}/MAX({rev_g_ref},0.001))"
                  f"*({vals['ebitda_margin']}/MAX({ebitda_m_ref},0.001))"
                  f"*(({wacc_ref}-{tgr_ref})/MAX({vals['wacc']}-{vals['tgr']},0.001))")
        c = ws.cell(row=row, column=7, value=approx)
        c.number_format = FMT_COMMA2
        c.font = Font(bold=True, size=10)
        c.fill = _fill(fill_map[name])
        row += 1

    # ── Key Value Drivers callout ────────────────────────────────────────────
    row += 2
    _header_row(ws, row, "KEY VALUE DRIVERS IDENTIFIED BY AI", 1, 6)
    row += 1
    for i, d in enumerate(drivers[:3]):
        ws.cell(row=row, column=1, value=f"Driver {i+1}: {d.get('driver','')}")
        ws.cell(row=row, column=1).font = Font(bold=True, size=10)
        ws.cell(row=row, column=2, value=d.get("description", ""))
        ws.cell(row=row, column=2).font = Font(size=10, italic=True)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        row += 1

    return ws


# ═════════════════════════════════════════════════════════════════════════════
# MAIN ENTRY POINT
# ═════════════════════════════════════════════════════════════════════════════
def build_consumer_retail_model(fin: dict, output_path: str):
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    assumptions_ws = _build_assumptions(wb, fin)
    is_ws          = _build_income_statement(wb, fin, assumptions_ws)
    cf_ws          = _build_cash_flow(wb, fin, assumptions_ws, is_ws)
    bs_ws          = _build_balance_sheet(wb, fin, assumptions_ws, is_ws)
    dcf_ws         = _build_dcf(wb, fin, assumptions_ws, is_ws, bs_ws, cf_ws)
    _build_sensitivity(wb, fin, assumptions_ws, dcf_ws)

    wb.save(output_path)
